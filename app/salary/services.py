from typing import Any
import re
import os
from decouple import config
from django.conf import settings
import openpyxl
from django.http import HttpResponse
from django.core.files.uploadedfile import File
from django.db.models import Prefetch

from jinja2 import Environment, FileSystemLoader
from utils.jinja_core import JinjaCore
from utils.mail_core import MailCore
from weasyprint import HTML
from PyPDF2 import PdfReader, PdfWriter
from pandas import DataFrame

from datetime import date, datetime 
from attendance.models import Attendance
from contract.models import Contract
from salary.constants import ENTERPRISE_PERCENT_HEALTH_INSURANCE, ENTERPRISE_PERCENT_SOCIAL_INSURANCE, ENTERPRISE_PERCENT_UNEMOLOYMENT_INSURANCE, EXPORT_FILE_NAME, HAVING_LUNCH, PERCENT_HEALTH_INSURANCE, PERCENT_SOCIAL_INSURANCE, PERCENT_UNEMOLOYMENT_INSURANCE, ROLE_MEMBER, SKIN, TAXABLE_EARNING_BASE_COST, TAXABLE_EARNING_UNIT_COST
from salary.models import Salary
from salary.utils import Utils
from user_core.models import CustomUser
from utils.datetime_core import DateTimeCore


class SalaryService:

    @staticmethod
    def get_date_attendance_and_contract_by_user(users: list[str], month: int, year: int) -> DataFrame:
        attendance_qs = Attendance.objects.filter(date__year=year, date__month=month)
        attendance_prefetch = Prefetch("attendance_user", queryset=attendance_qs, to_attr="filtered_attendance")

        contract_qs = Contract.objects.filter(is_current_contract=True)
        contract_prefetch = Prefetch("contract_user", queryset=contract_qs, to_attr="filtered_contract")

        users_qs = CustomUser.objects.filter(username__in=users).prefetch_related(attendance_prefetch, contract_prefetch)

        df = DataFrame([
            {
                "user": user,
                "attendance": user.filtered_attendance[0] if user.filtered_attendance else None,
                "contract": user.filtered_contract[0] if user.filtered_contract else None
            }
            for user in users_qs
        ])

        return df

    @staticmethod
    def create_bulk_salary_service(users: list[str], month: int, year: int, working_day: int, override: bool, 
                                        underpaid_overpaid: int, increased_income: int, income_deduction: int, other: int):
            if working_day <= 0:
                working_day = DateTimeCore.count_weekdays(year, month)  

            data_create = SalaryService.get_date_attendance_and_contract_by_user(users, month, year)

            if data_create.empty:
                return False

            data_create["salary_basic"] = data_create["contract"].apply(lambda x: x.salary_basic)
            data_create["working_days_month"] = data_create["attendance"].apply(lambda x: x.working_days_month)
            data_create["actual_ot"] = data_create["attendance"].apply(lambda x: x.actual_ot)
            data_create["coefficient_ot"] = data_create["attendance"].apply(lambda x: x.coefficient_ot)

            data_create["daily_wage"] = (data_create["salary_basic"] / working_day) * data_create["working_days_month"]
            data_create["skin"] = (SKIN / working_day) * data_create["working_days_month"]
            data_create["having_lunch"] = (HAVING_LUNCH / working_day) * data_create["working_days_month"]
            data_create["allowances_excluding_tax"] = data_create["skin"] + data_create["having_lunch"]

            data_create["allowances_taxable"] = data_create["contract"].apply(lambda x: (
                x.allowance_housing + x.allowance_phone + x.allowance_fuel + 
                x.allowance_responsibility + x.bonus_kpi
            ))

            data_create["overtime_tax"] = data_create["actual_ot"]
            data_create["overtime_no_tax"] = data_create["coefficient_ot"] - data_create["actual_ot"]
            data_create["overtime"] = data_create["overtime_tax"] + data_create["overtime_no_tax"]

            data_create["total_income"] = data_create["daily_wage"] + data_create["allowances_excluding_tax"] + data_create["allowances_taxable"] + data_create["overtime"]

            data_create["social_insurance"] = data_create["salary_basic"] * PERCENT_SOCIAL_INSURANCE
            data_create["health_insurance"] = data_create["salary_basic"] * PERCENT_HEALTH_INSURANCE
            data_create["unemployment_insurance"] = data_create["salary_basic"] * PERCENT_UNEMOLOYMENT_INSURANCE
            data_create["insurance"] = data_create["social_insurance"] + data_create["health_insurance"] + data_create["unemployment_insurance"]

            data_create["enterprise_insurance_social"] = data_create["salary_basic"] * ENTERPRISE_PERCENT_SOCIAL_INSURANCE
            data_create["enterprise_insurance_health"] = data_create["salary_basic"] * ENTERPRISE_PERCENT_HEALTH_INSURANCE
            data_create["enterprise_insurance_unemployment"] = data_create["salary_basic"] * ENTERPRISE_PERCENT_UNEMOLOYMENT_INSURANCE
            data_create["count_enterprise_insurance"] = data_create["enterprise_insurance_social"] + data_create["enterprise_insurance_health"] + data_create["enterprise_insurance_unemployment"]

            data_create["taxable_earning"] = data_create["daily_wage"] + data_create["allowances_taxable"] + data_create["overtime_tax"]
            data_create["dependent"] = data_create["contract"].apply(lambda x: x.number_dependents)
            data_create["taxable_income"] = (data_create["taxable_earning"] - data_create["insurance"] - TAXABLE_EARNING_BASE_COST - (data_create["dependent"] * TAXABLE_EARNING_UNIT_COST)).clip(lower=0)

            data_create["personal_income_tax"] = data_create["taxable_income"].apply(Utils.calculate_tax) + float(underpaid_overpaid)

            data_create["other_income_deductions"] = float(increased_income) - float(income_deduction) + float(other)
            data_create["actual_income"] = data_create["total_income"] - data_create["insurance"] - data_create["personal_income_tax"] + data_create["other_income_deductions"]

            format_date_pay_period = datetime.strptime(f"{year}-{month}-01", "%Y-%m-%d")
            data_create["date"] = format_date_pay_period

            existing_salaries = Salary.objects.filter(user__in=data_create["user"].tolist(), date__year=year, date__month=month, deleted_at__isnull=True)
            existing_users = set(existing_salaries.values_list("user", flat=True))

            if not override:
                data_create = data_create[~data_create["user"].isin(existing_users)]

            if data_create.empty:
                return True
            try:
            # Bulk update or create
                salary_records = [
                    Salary(
                        user=row["user"],
                        date=row["date"],
                        total_income=row["total_income"],
                        daily_wage=row["daily_wage"],
                        overtime_taxable=row["overtime_tax"],
                        overtime_non_taxable=row["overtime_no_tax"],
                        employee_insurance_social=row["social_insurance"],
                        employee_insurance_health=row["health_insurance"],
                        employee_insurance_unemployment=row["unemployment_insurance"],
                        count_employee_insurance=row["insurance"],
                        enterprise_insurance_social=row["enterprise_insurance_social"],
                        enterprise_insurance_health=row["enterprise_insurance_health"],
                        enterprise_insurance_unemployment=row["enterprise_insurance_unemployment"],
                        count_enterprise_insurance=row["count_enterprise_insurance"],
                        personal_income_tax=row["personal_income_tax"],
                        taxable_income=row["taxable_income"],
                        dependent=row["dependent"],
                        taxable_earnings=row["taxable_earning"],
                        tax_underpaid_or_overpaid=underpaid_overpaid,
                        income_additional=increased_income,
                        income_deduction=income_deduction,
                        other=other,
                        count_other_income=row["other_income_deductions"],
                        actual_income=row["actual_income"],
                        working_day=working_day
                    )
                    for _, row in data_create.iterrows()
                ]

                Salary.objects.bulk_create(salary_records, ignore_conflicts=True)


                for _, row in data_create.iterrows():
                    salary = Salary.objects.filter(user=row["user"], date__year=year, date__month=month).first()
                    SalaryService.flow_create_payslips(row["user"], salary, row["attendance"], row["contract"], month, year)

                return True

            except Exception:
                return False

    @staticmethod    
    def find_export_start(sheet):
        """Find the first row containing `{{}}`"""
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and re.search(r"{{(.*?)}}", cell.value):
                    return cell.row, cell.column
        return None, None  

    @staticmethod
    def export_salary_service(queryset=None, file: File = None, users: list[str] = None, month:int = None, year:int = None) -> HttpResponse:
        if file is None:
            
            default_file_path = os.path.join(settings.BASE_DIR, 'app', 'salary', 'templates', 'admin', 'template-export', 'template-export-salary.xlsx')
            with open(default_file_path, 'rb') as default_file:
                wb = openpyxl.load_workbook(default_file)
        else:
            wb = openpyxl.load_workbook(file)
        sheet = wb.active

        start_row, start_col = SalaryService.find_export_start(sheet)
        if not start_row or not start_col:
            return HttpResponse("Template {{}} was not found in the Excel file.", status=400)

        
        template_var_positions = {}  
        for col in range(start_col, sheet.max_column + 1):
            cell_value = sheet.cell(row=start_row, column=col).value 
            match = re.search(r"{{(.*?)}}", str(cell_value) if cell_value else "")
            if match:
                template_var_positions[match.group(1)] = col

        fields = [
            "user__first_name", "user__last_name", "date", "user__phone", "total_income", "daily_wage", 
            "overtime_taxable", "overtime_non_taxable", "employee_insurance_social", "employee_insurance_health", 
            "employee_insurance_unemployment", "count_employee_insurance", "enterprise_insurance_social", 
            "enterprise_insurance_health", "enterprise_insurance_unemployment", "count_enterprise_insurance", 
            "personal_income_tax", "taxable_income", "dependent", "taxable_earnings", "tax_underpaid_or_overpaid", 
            "income_additional", "income_deduction", "other", "count_other_income", "actual_income", "working_day", 
            "user__username"
        ]

        if queryset is None:
            salaries = Salary.objects.filter(
                user__username__in=users, 
                date__year=year,
                date__month=month
            ).values_list(*fields)
        else:
            salaries = queryset.values_list(*fields)

        row_idx = start_row
        for row_data in salaries:
            data = dict(zip(fields, row_data))

            column_map = {
                "user": f"{data['user__last_name']} {data['user__first_name']}".strip(),
                "date": data["date"].strftime("%m/%Y") if data["date"] else "",
                "role": ROLE_MEMBER,
                "phone": data["user__phone"],
                **{k: f"{data[k]:,}" for k in fields if k not in ["user__first_name", "user__last_name", "date", "user__phone", "user__username", "working_day"]},
                "working_day": data["working_day"],
                "overtime": data["overtime_taxable"] + data["overtime_non_taxable"]
            }

            get_user = CustomUser.objects.filter(username=data["user__username"]).first()
            
            contract = Contract.objects.filter(user=get_user, is_current_contract=True).first()
            
            attendance = Attendance.objects.filter(user=get_user, date=data["date"]).first()

            if contract:

                allowance_housing = contract.allowance_housing

                allowance_phone = contract.allowance_phone

                allowance_fuel = contract.allowance_fuel

                allowance_responsibility = contract.allowance_responsibility

                bonus_kpi = contract.bonus_kpi

                allowances_taxable = allowance_housing + allowance_phone + allowance_fuel + allowance_responsibility + bonus_kpi
                
                column_map.update({
                    "salary_basic": f"{contract.salary_basic:,}",
                    "allowance_housing": allowance_housing,
                    "allowance_phone": allowance_phone,
                    "allowance_fuel": allowance_fuel,
                    "allowance_responsibility": allowance_responsibility,
                    "bonus_kpi": bonus_kpi,
                    "allowances_taxable": allowances_taxable
                })

            if attendance:
                working_days_month = attendance.working_days_month
                working_day = data["working_day"]
                if working_day != 0.0 and working_days_month != 0.0:
                    skin = SKIN / working_day * working_days_month
                    having_lunch = HAVING_LUNCH / working_day * working_days_month
                else:
                    skin = 0
                    having_lunch = 0
                
                column_map.update({
                    "working_days_month": working_days_month,
                    "actual_ot": attendance.actual_ot,
                    "coefficient_ot": attendance.coefficient_ot,
                    "leave": f"{attendance.take_leave_month} / {attendance.maximum_leave - attendance.leave_taken} / {attendance.maximum_leave}",
                    "skin": skin,
                    "having_lunch": having_lunch,
                    "allowances_excluding_tax": skin + having_lunch
                })

            for var_name, col_idx in template_var_positions.items():
                if var_name in column_map:
                    sheet.cell(row=row_idx, column=col_idx, value=column_map[var_name])

            row_idx += 1

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename={EXPORT_FILE_NAME}"
        wb.save(response)
        return response

    @staticmethod
    def convert_salary_dict(user: CustomUser, salary: Salary, attendance: Attendance, contract: Contract) -> dict:

        allowances = {
            "allowance_housing": contract.allowance_housing,
            "allowance_phone": contract.allowance_phone,
            "allowance_fuel": contract.allowance_fuel,
            "allowance_responsibility": contract.allowance_responsibility,
            "bonus_kpi": contract.bonus_kpi,
        }
        allowances_taxable = sum(allowances.values())

        working_days_month = attendance.working_days_month

        working_day = salary.working_day

        if working_day != 0.0 and working_days_month != 0.0:
            skin = SKIN / working_day * working_days_month
            having_lunch = HAVING_LUNCH / working_day * working_days_month
        else:
            skin = 0
            having_lunch = 0

        return {
            "full_name": f"{user.last_name} {user.first_name}",
            "user_name": user.username,
            "user_phone": user.phone,
            "date_of_birth": user.date_of_birth,
            "bank_number_id": user.bank_number_id,
            "bank_name": user.bank_name,
            "bank_beneficiary": user.bank_beneficiary,
            "date": salary.date,
            "role": ROLE_MEMBER,
            "salary_basic": contract.salary_basic,
            "working_days_month": working_days_month,
            "actual_ot": attendance.actual_ot,
            "coefficient_ot": attendance.coefficient_ot,
            "take_leave_month": attendance.take_leave_month,
            "number_days_off": attendance.maximum_leave - attendance.leave_taken,
            "maximum_leave": attendance.maximum_leave,
            "total_income": salary.total_income,
            "daily_wage": salary.daily_wage,
            "allowances_excluding_tax": skin + having_lunch,
            "skin": skin,
            "having_lunch": having_lunch,
            "allowances_taxable": allowances_taxable,
            **allowances,
            "overtime": salary.overtime_taxable + salary.overtime_non_taxable,
            "overtime_taxable": salary.overtime_taxable,
            "overtime_non_taxable": salary.overtime_non_taxable,
            "employee_insurance_social": salary.employee_insurance_social,
            "employee_insurance_health": salary.employee_insurance_health,
            "employee_insurance_unemployment": salary.employee_insurance_unemployment,
            "count_employee_insurance": salary.count_employee_insurance,
            "personal_income_tax": salary.personal_income_tax,
            "taxable_income": salary.taxable_income,
            "dependent": salary.dependent,
            "taxable_earnings": salary.taxable_earnings,
            "tax_underpaid_or_overpaid": salary.tax_underpaid_or_overpaid,
            "income_additional": salary.income_additional,
            "income_deduction": salary.income_deduction,
            "other": salary.other,
            "count_other_income": salary.count_other_income,
            "actual_income": salary.actual_income,
            "working_day": working_day
        }
    
    @staticmethod
    def flow_create_payslips(user, salary, attendance, contract, month: int, year: int):

        # Create payslip file
        month_after = DateTimeCore.add_month(year, month, 1)
        deadline = DateTimeCore.get_date_text(
            month_after.year, month_after.month, 10, '%d/%m/%Y')

        payslip_dict = SalaryService.convert_salary_dict(user, salary, attendance, contract)
        SalaryService.render_payslip(payslip_dict, deadline=deadline)

            
    @staticmethod
    def render_payslip(payslip: dict[str, Any], deadline: str, template='templates.html') -> str:
        
        template_dir = os.path.join(os.path.dirname(__file__), "templates", "payslip")
        env = Environment(loader=FileSystemLoader(template_dir))
        env.filters['vnd'] = JinjaCore.viet_nam_dong_format
        env.filters['int_priority'] = JinjaCore.integer_priority_format
        static_dir = os.path.join(os.path.dirname(__file__), "static")
        template_file = env.get_template(template)
        image_path = os.path.join(static_dir, 'images', 'logo.png')
        absolute_image_path = os.path.abspath(image_path)

        date = payslip['date']
        file_name = date.strftime('%m-%Y')
        payslip['date'] = date.strftime('%m/%Y')
        
        output = template_file.render(
            data=payslip, logo=absolute_image_path, deadline=deadline)

        output_pdf_path = f"./app/salary/static/payslips_data/{
            payslip["user_name"]}/{file_name}.pdf"
        password_protected_pdf_path = f"./app/salary/static/payslips_data/{
            payslip["user_name"]}/{file_name}-protect.pdf"

        folder_path = os.path.dirname(output_pdf_path)
        # Create the folder if it doesn't exist
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        password = payslip['date_of_birth'].strftime('%d%m%Y')    

        HTML(string=output).write_pdf(output_pdf_path)
        reader = PdfReader(output_pdf_path)
        writer = PdfWriter()

        for page in reader.pages:
            writer.add_page(page)

        writer.encrypt(password)
        with open(password_protected_pdf_path, "wb") as f:
            writer.write(f)

        SalaryService.upload_file_path_payslip(output_pdf_path, password_protected_pdf_path, payslip['user_name'], date)    
        return password_protected_pdf_path
    
    @staticmethod
    def upload_file_path_payslip(file_path: str, file_path_protect: str, user_name: str, date: date):

        get_user = CustomUser.objects.filter(username=user_name).first()

        Salary.objects.filter(user=get_user, date=date).update(file_payslip=file_path, file_payslip_protect=file_path_protect)

    @staticmethod
    def handle_sendmail_payslip(queryset) -> bool:
        mail_client = MailCore(
            mail_server=config('LARK_MAIL_SERVER'),
            mail_port=int(config('LARK_MAIL_PORT')),
            mail_user=config('LARK_MAIL_SENDER'),
            mail_password=config('LARK_MAIL_PASSWORD')
        )
        try:
            for salary in queryset:
                user = salary.user
                payslip = {
                    "date_pay_period": salary.date.strftime('%m/%Y'),
                    "full_name": f"{user.last_name} {user.first_name}",
                    "email_member": user.business_email
                }
                Utils.send_payslips(mail_client, payslip, salary.file_payslip_protect)
            return True    
        except Exception:
            return False        
           
